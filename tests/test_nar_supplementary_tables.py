#!/usr/bin/env python3
"""Validate the submission-facing supplementary-table workbook."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

import openpyxl


EXPECTED_SHEETS = [
    "README",
    "S1_catalog_runs",
    "S2_study_summary",
    "S3_layer_coverage",
    "S4_analysis_inventory",
    "S5_null_calibration",
    "S6_PTPRC_evidence",
    "S7_MS4A1_usage",
    "S8_mouse_examples",
    "S9_mouse_scope",
    "S10_resource_comparison",
    "S11_release_completeness",
    "S12_release_contract",
    "S13_molecular_inventory",
    "S14_CD45_isoform_recovery",
    "S15_junction_program_effects",
    "Provenance",
]

# The frozen release. Cells are per study and are never summed over run rows:
# the run sum is 1,045,231 against the true 923,389.
RUNS, STUDIES, CELLS, EXCLUDED = 453, 34, 923_389, 35


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def records(sheet) -> list[dict[str, object]]:
    rows = list(sheet.values)
    return [dict(zip(rows[0], row)) for row in rows[1:]]


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    args = parser.parse_args()
    workbook = openpyxl.load_workbook(args.workbook, read_only=True, data_only=True)
    require(workbook.sheetnames == EXPECTED_SHEETS, "Unexpected workbook sheets")
    require(
        workbook["S1_catalog_runs"].max_row == RUNS + 1,
        f"S1 must list the {RUNS} released runs and nothing else",
    )
    require(
        workbook["S2_study_summary"].max_row == STUDIES + 1,
        f"S2 must have {STUDIES} data rows",
    )
    require(
        workbook["S10_resource_comparison"].max_row == 9,
        "S10 must compare scTHREAD with seven adjacent resources",
    )
    require(
        workbook["S11_release_completeness"].max_row == 3,
        "S11 must contain two audited completeness tiers",
    )
    require(
        workbook["S12_release_contract"].max_row == RUNS + 1,
        f"S12 must contain the {RUNS}-run availability contract",
    )
    require(
        workbook["S13_molecular_inventory"].max_row == 156,
        "S13 must contain the 155-run molecular recount",
    )
    require(
        "feature_fdr" not in next(workbook["S8_mouse_examples"].values),
        "S8 must not propagate pseudo-replicate FDR as formal evidence",
    )
    mouse_scope_columns = set(next(workbook["S9_mouse_scope"].values))
    require(
        not mouse_scope_columns.intersection(
            {"feature_fdr", "sig_genes", "sig_isoforms", "pval", "qval"}
        ),
        "S9 must be an evidence-boundary ledger, not pseudo-replicate inference",
    )

    runs = records(workbook["S1_catalog_runs"])
    require(
        len({row["srr"] for row in runs}) == RUNS,
        "S1 run IDs must be unique",
    )
    included = runs

    # The cell denominator lives in S2 and is per study. Summing S1's per-run
    # isoquant_cells gives 1,045,231, which is the error this asserts against.
    studies = records(workbook["S2_study_summary"])
    require(len(studies) == STUDIES, f"S2 must summarise {STUDIES} studies")
    require(
        sum(int(row["cells"]) for row in studies) == CELLS,
        f"S2 cell total must be {CELLS:,}",
    )
    require(
        {row["gse"] for row in studies} == {row["gse"] for row in included},
        "S2 must summarise exactly the studies S1 marks included",
    )
    require(
        sum(int(row["n_runs"]) for row in studies) == RUNS,
        f"S2 run counts must add to {RUNS}",
    )
    require(
        "isoquant_cells" not in studies[0],
        "S2 must not carry a run-summed cell column beside the authoritative one",
    )

    inventory = {
        row["axis"]: row for row in records(workbook["S4_analysis_inventory"])
    }
    source_path = (
        Path(__file__).resolve().parents[1] / "tables/Table_S_three_axis_summary.tsv"
    )
    with source_path.open(newline="") as handle:
        source_inventory = {
            row["axis"]: row for row in csv.DictReader(handle, delimiter="\t")
        }
    for axis in ("ASE", "DIU", "APA"):
        observed = (
            int(inventory[axis]["n_genes_tested"]),
            int(inventory[axis]["n_sig_fdr05_effect_gate"]),
        )
        expected = (
            int(source_inventory[axis]["n_genes_tested"]),
            int(source_inventory[axis]["n_sig_fdr05_effect_gate"]),
        )
        require(observed == expected, f"{axis} inventory mismatch")
        require(
            str(source_inventory[axis]["source"]).endswith("_observed_9999.tsv"),
            f"{axis} is not linked to the 9,999-permutation source",
        )

    calibration = records(workbook["S5_null_calibration"])
    null_rows = [row for row in calibration if row["dataset"] == "null"]
    observed_rows = {
        row["analysis"]: row
        for row in calibration
        if row["dataset"] == "observed"
    }
    require(len(null_rows) == 9, "Expected nine complete null rows")
    require(all(int(row["significant"]) == 0 for row in null_rows), "Null discovery")
    require(
        {
            axis: (int(row["rows"]), int(row["significant"]))
            for axis, row in observed_rows.items()
        }
        == {"ase": (6930, 0), "diu": (8092, 2008), "apa": (10531, 2558)},
        "S5 observed rows are not synchronized to the 9,999-permutation results",
    )
    require(
        all(int(row["inner_permutations"]) == 9999 for row in observed_rows.values())
        and all(int(row["inner_permutations"]) == 200 for row in null_rows),
        "S5 inner-permutation counts are missing or incorrect",
    )
    require(
        all(not str(row["path"]).startswith("/") for row in calibration),
        "Internal absolute paths leaked",
    )
    resources = {
        row["resource"] for row in records(workbook["S10_resource_comparison"])
    }
    require(
        resources
        == {
            "scTHREAD",
            "FLIBase",
            "ISOdb",
            "isoformAtlas",
            "scAPAdb",
            "PolyASite v3.0",
            "PolyA_DB v4",
            "RMPore",
        },
        "Resource comparison is incomplete",
    )
    tiers = {
        row["tier"]: row
        for row in records(workbook["S11_release_completeness"])
    }
    require(
        int(tiers["transcript_count_documented"]["n_runs"]) == 155,
        "Transcript-count-documented tier changed",
    )
    require(
        int(tiers["one_well_membership_only"]["n_runs"]) == 298
        and int(tiers["one_well_membership_only"]["n_pipeline_called_cells"]) == 298,
        "Membership-only tier changed",
    )
    require(
        sum(int(row["n_runs"]) for row in tiers.values()) == RUNS,
        f"the completeness tiers must partition all {RUNS} runs",
    )
    # a run-level cell sum in the same workbook as the per-study denominator is
    # how a second, smaller total gets quoted as the release size
    require(
        "n_cells" not in next(iter(tiers.values())),
        "S11 must name its cell column as a run-level pipeline count",
    )
    require(
        tiers["one_well_membership_only"]["pipeline_values"]
        == "EXCLUDE|unclassified"
        and tiers["one_well_membership_only"]["cell_count_methods"]
        == "smartseq2_one_well",
        "Membership-only definition changed",
    )
    contract = records(workbook["S12_release_contract"])
    statuses = {}
    for row in contract:
        status = row["file_contract_status"]
        statuses[status] = statuses.get(status, 0) + 1
    require(
        statuses == {"isoquant_core_complete": 155, "membership_only": 298}
        and sum(statuses.values()) == RUNS,
        "S12 feature-contract tiers changed",
    )
    molecular = records(workbook["S13_molecular_inventory"])
    require(
        all(
            int(row["observed_discovered_run_transcript_detections"])
            == int(row["observed_discovered_reference_transcript_detections"])
            + int(row["observed_discovered_isoquant_novel_transcript_detections"])
            + int(row["observed_discovered_other_transcript_detections"])
            for row in molecular
        ),
        "S13 discovered-transcript class totals do not reconcile",
    )
    mismatches = {
        row["srr"] for row in molecular if not row["registry_recount_match"]
    }
    # the registry once disagreed with the files for SRR33259867; on the frozen
    # release every run reconciles, so any mismatch is a regression
    require(
        not mismatches,
        f"S13 registry-versus-file recount disagrees for {sorted(mismatches)}",
    )
    for sheet_name in ("S12_release_contract", "S13_molecular_inventory"):
        require(
            all(
                not (isinstance(value, str) and value.startswith("/gpfs/"))
                for row in workbook[sheet_name].iter_rows(values_only=True)
                for value in row
            ),
            f"{sheet_name} leaks internal absolute paths",
        )
    print("SUPPLEMENTARY TABLE WORKBOOK QA PASS")


if __name__ == "__main__":
    main()
