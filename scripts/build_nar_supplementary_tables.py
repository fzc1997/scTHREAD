#!/usr/bin/env python3
"""Build the submission-facing scTHREAD supplementary-table workbook."""

from __future__ import annotations

import argparse
import hashlib
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
EXCLUDED_RUNS = ROOT / "tables/modality_audit/excluded_runs_ledger.tsv"
RELEASE = ROOT / "tables/release_20260803"
# Cells are authoritative per study and must never be summed over run rows:
# the run sum is 1,045,231 against the true 923,389, because 14 studies report
# a study-level count that the per-run caller cannot reproduce.
STUDY_CELLS = RELEASE / "release_study_cells.tsv"
# The frozen manifest leaves biology_group and description empty, and the atlas
# fills them in a mix of Chinese and English. tissue and system_display are the
# only fully English biology fields, so they are what the workbook ships.
STUDY_ATLAS = ROOT / "tables/fig1_study_atlas.tsv"
SOURCES = {
    "S1_catalog_runs": RELEASE / "release_manifest.tsv",
    "S3_layer_coverage": ROOT / "tables/fig1_layer_coverage_labeled.tsv",
    "S4_analysis_inventory": ROOT / "tables/Table_S_three_axis_summary.tsv",
    "S5_null_calibration": ROOT
    / "tables/p0_biological_unit_rerun/validation_summary.tsv",
    "S6_PTPRC_evidence": ROOT / "tables/Table_S_PTPRC_multievidence.tsv",
    "S7_MS4A1_usage": ROOT / "tables/MS4A1_isoform_usage_by_ct.tsv",
    "S8_mouse_examples": ROOT / "tables/Fig3_mouse_scONT_portal_DTU_examples.tsv",
    "S9_mouse_scope": ROOT / "tables/Table_S_mouse_evidence_boundaries.tsv",
    "S10_resource_comparison": ROOT / "tables/Table_S_resource_comparison.tsv",
    "S11_release_completeness": ROOT
    / "tables/Table_S_release_completeness_tiers.tsv",
    "S12_release_contract": ROOT
    / "tables/release_content/run_layer_contract.tsv",
    "S13_molecular_inventory": ROOT
    / "tables/release_content/run_molecular_inventory.tsv",
    "S14_CD45_isoform_recovery": ROOT / "tables/cd45_ra_ro_recovery.tsv",
    "S15_junction_program_effects": ROOT
    / "tables/fig4_source/junction_program_effects_all_genes.tsv",
}

DESCRIPTIONS = {
    "S1_catalog_runs": (
        "One row per released sequencing run. A further 35 candidate runs were "
        "excluded as non-transcriptomic libraries and are not listed."
    ),
    "S2_study_summary": (
        "Study-level summary of the frozen 453-run release. Cell counts are the "
        "authoritative per-study values, not sums over run rows."
    ),
    "S3_layer_coverage": (
        "Evidence-layer run coverage across the 15 studies that carry "
        "layer-level evidence."
    ),
    "S4_analysis_inventory": (
        "Biological-unit ASE, DIU and APA results from 9,999 restricted permutations."
    ),
    "S5_null_calibration": (
        "Observed and three-seed complete outer-null validation ledger."
    ),
    "S6_PTPRC_evidence": (
        "Corrected multi-layer statistics and portal/API routes for PTPRC."
    ),
    "S7_MS4A1_usage": (
        "Descriptive MS4A1 isoform-usage fractions by cell type."
    ),
    "S8_mouse_examples": (
        "Selected previously published mouse scONT portal and descriptive "
        "usage examples; pseudo-replicate FDR fields are omitted."
    ),
    "S9_mouse_scope": (
        "Evidence boundaries for the previously published mouse scONT use case."
    ),
    "S10_resource_comparison": (
        "Feature-by-feature positioning against adjacent isoform, APA and "
        "single-molecule resources, with primary references."
    ),
    "S11_release_completeness": (
        "Records carrying run-level transcript counts in the frozen "
        "453-run manifest."
    ),
    "S12_release_contract": (
        "Run-by-layer availability contract for the frozen 453-run manuscript "
        "snapshot; ASE is represented by availability metadata only."
    ),
    "S13_molecular_inventory": (
        "Run-level recount of positive reference and discovered IsoQuant "
        "transcript and gene records in the 155 feature-complete runs."
    ),
    "S14_CD45_isoform_recovery": (
        "Molecule counts by cell type for the two annotated PTPRC models that "
        "differ by exactly the three variable exons defining CD45RO."
    ),
    "S15_junction_program_effects": (
        "Per-biological-unit monocyte-minus-T differences in within-gene "
        "exact-junction usage for the five loci that passed correction, in "
        "every cohort in which each was measured."
    ),
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def relative_source(value: object) -> object:
    """Record a source as a repository-relative path.

    Falling back to the bare filename loses the provenance the sheet exists to
    carry, and it happens whenever the workbook is rebuilt from a checkout at a
    different location than the one that produced the recorded path. Anchor on
    the first known top-level directory instead, so the recorded source reads
    the same wherever it was built.
    """
    if not isinstance(value, str) or not value.startswith("/"):
        return value
    path = Path(value)
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        parts = path.parts
        for anchor in ("tables", "results", "figures", "manuscript"):
            if anchor in parts:
                return str(Path(*parts[parts.index(anchor):]))
        return path.name


def load_tables() -> dict[str, pd.DataFrame]:
    tables = {
        name: pd.read_csv(path, sep="\t", keep_default_na=False)
        for name, path in SOURCES.items()
    }
    tables["S8_mouse_examples"] = tables["S8_mouse_examples"].drop(
        columns=["feature_fdr"], errors="ignore"
    )
    # S1 is the released run list. The 35 excluded candidates were carried here
    # for a while so a reader could check the scope rule, but every one of them
    # fails the same test - a non-transcriptomic INSDC library source - and the
    # Methods states that rule, so the rows added length without information.
    runs = tables["S1_catalog_runs"].copy()
    excluded = pd.read_csv(EXCLUDED_RUNS, sep="\t", keep_default_na=False)
    atlas = pd.read_csv(STUDY_ATLAS, sep="\t", keep_default_na=False)
    biology = atlas[["gse", "tissue", "system_display"]].rename(
        columns={"system_display": "biological_system"}
    )
    ledger = runs.fillna("")
    ledger = ledger.drop(columns=["biology_group", "description"], errors="ignore")
    ledger = ledger.merge(biology, on="gse", how="left").fillna("")
    tables["S1_catalog_runs"] = ledger

    included = ledger
    study_cells = pd.read_csv(STUDY_CELLS, sep="\t", keep_default_na=False)
    if len(included) != 453:
        raise RuntimeError(f"frozen release is 453 runs, S1 lists {len(included)}")
    if len(excluded) != 35:
        raise RuntimeError(f"exclusion ledger should hold 35 runs, has {len(excluded)}")
    if set(included["gse"]) != set(study_cells["gse"]):
        raise RuntimeError("manifest studies and the study-cell authority disagree")
    total_cells = int(pd.to_numeric(study_cells["cells"]).sum())
    if len(study_cells) != 34 or total_cells != 923_389:
        raise RuntimeError(
            f"frozen release is 34 studies / 923,389 cells, authority gives "
            f"{len(study_cells)} / {total_cells:,}"
        )
    runs = included
    studies = (
        runs.groupby("gse", as_index=False, dropna=False)
        .agg(
            species=("species", lambda values: "|".join(sorted(set(values)))),
            platforms=("platform", lambda values: "|".join(sorted(set(values)))),
            n_runs=("srr", "nunique"),
        )
        .merge(
            study_cells[["gse", "cells", "cell_count_method"]], on="gse", how="left"
        )
        .merge(biology, on="gse", how="left")
        .sort_values("gse")
        .reset_index(drop=True)
    )
    if len(studies) != 34 or int(studies["cells"].sum()) != 923_389:
        raise RuntimeError("Frozen study summary must be 34 studies / 923,389 cells")
    tables = {
        "S1_catalog_runs": tables["S1_catalog_runs"],
        "S2_study_summary": studies,
        **{key: value for key, value in tables.items() if key != "S1_catalog_runs"},
    }
    for name in ("S4_analysis_inventory", "S5_null_calibration"):
        for column in ("source", "path"):
            if column in tables[name]:
                tables[name][column] = tables[name][column].map(relative_source)
    # S12 still carries the column name from when the snapshot was a candidate,
    # while holding the frozen release id; S1 calls the same field "release"
    tables["S12_release_contract"] = tables["S12_release_contract"].rename(
        columns={"release_candidate": "release"}
    )
    for name in ("S12_release_contract", "S13_molecular_inventory"):
        for column in tables[name].columns:
            if "path" in column or column.endswith("_dir"):
                tables[name][column] = tables[name][column].map(relative_source)
    return tables


def readme_frame(tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = [
        {
            "item": "Workbook",
            "description": "scTHREAD NAR Database Issue supplementary tables",
            "value": "release_20260803",
        },
        {
            "item": "Frozen denominator",
            "description": "Runs / studies / cells",
            "value": "453 / 34 / 923,389",
        },
        {
            "item": "Scope boundary",
            "description": (
                "The rolling web registry is not the manuscript denominator."
            ),
            "value": "Use S1 and S2 for manuscript release counts",
        },
        {
            "item": "Mouse inference boundary",
            "description": (
                "Mouse DTU uses cell-bootstrap pseudo-replicates, not independent embryos."
            ),
            "value": "Utility demonstration only",
        },
    ]
    rows.extend(
        {
            "item": name,
            "description": DESCRIPTIONS[name],
            "value": f"{len(frame):,} data rows",
        }
        for name, frame in tables.items()
    )
    return pd.DataFrame(rows)


def provenance_frame(tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    for name, frame in tables.items():
        if name == "S2_study_summary":
            rows.append(
                {
                    "sheet": name,
                    "source_file": "derived from S1_catalog_runs",
                    "source_sha256": sha256(SOURCES["S1_catalog_runs"]),
                    "rows": len(frame),
                    "derivation": (
                        "group by gse for runs and platforms; cells joined from "
                        "release_study_cells.tsv, never summed over run rows"
                    ),
                }
            )
            continue
        path = SOURCES[name]
        rows.append(
            {
                "sheet": name,
                "source_file": str(path.relative_to(ROOT)),
                "source_sha256": sha256(path),
                "rows": len(frame),
                "derivation": "verbatim rows; absolute source paths shortened"
                if name in {"S4_analysis_inventory", "S5_null_calibration"}
                else "verbatim",
            }
        )
    return pd.DataFrame(rows)


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="DCEFEF")
    for worksheet in writer.book.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="17324D")
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column_index, cells in enumerate(worksheet.columns, start=1):
            values = ["" if cell.value is None else str(cell.value) for cell in cells]
            width = min(max(max(map(len, values), default=0) + 2, 10), 45)
            worksheet.column_dimensions[get_column_letter(column_index)].width = width
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)


def build(output: Path) -> None:
    tables = load_tables()
    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        readme_frame(tables).to_excel(writer, sheet_name="README", index=False)
        for name, frame in tables.items():
            frame.to_excel(writer, sheet_name=name, index=False)
        provenance_frame(tables).to_excel(
            writer, sheet_name="Provenance", index=False
        )
        style_workbook(writer)
    print(f"{output}\t{output.stat().st_size} bytes\tsha256={sha256(output)}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        type=Path,
        default=ROOT
        / "manuscript/supplementary/scTHREAD_Supplementary_Tables.xlsx",
    )
    args = parser.parse_args()
    build(args.output.resolve())


if __name__ == "__main__":
    main()
